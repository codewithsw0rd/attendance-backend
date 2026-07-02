from .serializers import *
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.exceptions import PermissionDenied
from django.http import HttpResponse
from ..models import Subject, Enrollment, ClassSession
from ..filters import SubjectFilter, EnrollmentFilter, ClassSessionFilter
from core.utils.custom_perms import IsClientUser, IsAdminOrTeacher
from core.utils.sort import apply_sorting
from accounts.models import UserType
import openpyxl
from io import BytesIO


class SubjectViewSet(viewsets.ModelViewSet):
    queryset = Subject.objects.all()
    serializer_class = SubjectSerializer
    filterset_class = SubjectFilter
    search_fields = ['name', 'code', 'department']
    
    # Sorting configuration
    ordering_fields = ['id', 'name', 'code', 'department', 'semester', 'teacher__first_name', 'created_at', 'updated_at']
    ordering = ['created_at']
    SORT_MAPPING = {
        'id': 'id',
        'name': 'name',
        'code': 'code',
        'department': 'department',
        'semester': 'semester',
        'teacher_name': 'teacher__first_name',
        'created_at': 'created_at',
        'updated_at': 'updated_at',
    }

    def get_permissions(self):
        if self.action in ('create', 'update', 'partial_update', 'destroy'):
            return [IsClientUser()]
        return [IsAdminOrTeacher()]

    def get_serializer_class(self):
        """Use read serializer for list/retrieve, write serializer for create/update"""
        if self.action in ['list', 'retrieve']:
            return SubjectReadSerializer
        return SubjectSerializer

    def get_queryset(self):
        user = self.request.user
        if not user.is_authenticated:
            return Subject.objects.none()
        if user.user_type == UserType.ADMIN:
            return Subject.objects.all()
        if user.user_type == UserType.TEACHER:
            return Subject.objects.filter(teacher__user=user)
        return Subject.objects.none()
    
    def list(self, request, *args, **kwargs):
        self.queryset = apply_sorting(request, self.get_queryset(), self)
        return super().list(request, *args, **kwargs)
    
    @action(detail=False, methods=['get'], url_path='export_excel')
    def export_excel(self, request):
        """
        Export subjects to Excel file with filtering, searching, sorting support.
        Query parameters:
            - search: Search by name, code, department
            - department: Filter by department
            - semester: Filter by semester
            - ordering: Sort field (e.g., 'name', '-created_at')
        """
        # Apply all the same filters/search/ordering as the list view
        queryset = self.get_queryset()
        
        # Apply filters
        filterset = self.filterset_class(request.GET, queryset=queryset)
        queryset = filterset.qs
        
        # Apply search
        search_query = request.query_params.get('search', '')
        if search_query:
            from django.db.models import Q
            queryset = queryset.filter(
                Q(name__icontains=search_query) |
                Q(code__icontains=search_query) |
                Q(department__icontains=search_query)
            )
        
        # Apply ordering
        ordering = request.query_params.get('ordering', 'name')
        if ordering:
            queryset = queryset.order_by(ordering)
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Subjects"
        
        # Add headers
        headers = ['Subject Name', 'Code', 'Department', 'Semester', 'Teacher']
        ws.append(headers)
        
        # Style header row
        from openpyxl.styles import Font, PatternFill
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Add data rows
        for subject in queryset:
            teacher_name = ""
            if subject.teacher:
                teacher_name = f"{subject.teacher.user.first_name} {subject.teacher.user.last_name}".strip()
            
            ws.append([
                subject.name,
                subject.code,
                subject.department or "",
                subject.semester or "",
                teacher_name,
            ])
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 25
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Return as file download
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="subjects_{__import__("datetime").datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        return response
    
class EnrollmentViewSet(viewsets.ModelViewSet):
    queryset = Enrollment.objects.all()
    serializer_class = EnrollmentSerializer
    permission_classes = [IsClientUser]
    filterset_class = EnrollmentFilter
    search_fields = ['student__roll_number', 'subject__name', 'subject__code']
    
    # Sorting configuration
    ordering_fields = ['id', 'student__roll_number', 'student__department', 'student__year', 'subject__name', 'subject__code', 'created_at', 'updated_at']
    ordering = ['created_at']
    SORT_MAPPING = {
        'id': 'id',
        'student_id': 'student__id',
        'student_roll_number': 'student__roll_number',
        'student_department': 'student__department',
        'student_year': 'student__year',
        'subject_id': 'subject__id',
        'subject_name': 'subject__name',
        'subject_code': 'subject__code',
        'created_at': 'created_at',
        'updated_at': 'updated_at',
    }
    
    def get_serializer_class(self):
        """Use read serializer for list/retrieve, write serializer for create/update"""
        if self.action in ['list', 'retrieve']:
            return EnrollmentReadSerializer
        return EnrollmentSerializer
    
    def list(self, request, *args, **kwargs):
        self.queryset = apply_sorting(request, self.get_queryset(), self)
        return super().list(request, *args, **kwargs)
    
    @action(detail=False, methods=['get'], url_path='export_excel')
    def export_excel(self, request):
        """Export enrollments to Excel with filtering, searching, sorting"""
        queryset = self.get_queryset()
        
        filterset = self.filterset_class(request.GET, queryset=queryset)
        queryset = filterset.qs
        
        search_query = request.query_params.get('search', '')
        if search_query:
            from django.db.models import Q
            queryset = queryset.filter(
                Q(student__roll_number__icontains=search_query) |
                Q(subject__name__icontains=search_query) |
                Q(subject__code__icontains=search_query)
            )
        
        ordering = request.query_params.get('ordering', 'created_at')
        if ordering:
            queryset = queryset.order_by(ordering)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Enrollments"
        
        headers = ['Student Name', 'Roll No', 'Department', 'Subject', 'Code', 'Enrolled Date']
        ws.append(headers)
        
        from openpyxl.styles import Font, PatternFill
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        for enrollment in queryset:
            student_name = f"{enrollment.student.user.first_name} {enrollment.student.user.last_name}".strip()
            ws.append([
                student_name,
                enrollment.student.roll_number,
                enrollment.student.department or "",
                enrollment.subject.name,
                enrollment.subject.code,
                enrollment.created_at.strftime('%Y-%m-%d') if enrollment.created_at else "",
            ])
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="enrollments_{__import__("datetime").datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        return response
    
class ClassSessionViewSet(viewsets.ModelViewSet):
    queryset = ClassSession.objects.all()
    serializer_class = ClassSessionSerializer
    filterset_class = ClassSessionFilter
    search_fields = ['class_name', 'subject__name', 'subject__code']
    
    # Sorting configuration
    ordering_fields = ['id', 'subject__name', 'subject__code', 'class_name', 'date', 'start_time', 'end_time', 'created_at', 'updated_at']
    ordering = ['-date', 'start_time']
    SORT_MAPPING = {
        'id': 'id',
        'subject_id': 'subject__id',
        'subject_name': 'subject__name',
        'subject_code': 'subject__code',
        'class_name': 'class_name',
        'date': 'date',
        'start_time': 'start_time',
        'end_time': 'end_time',
        'created_at': 'created_at',
        'updated_at': 'updated_at',
    }

    def get_permissions(self):
        if self.action == 'destroy':
            return [IsClientUser()]
        return [IsAdminOrTeacher()]

    def get_serializer_class(self):
        """Use read serializer for list/retrieve, write serializer for create/update"""
        if self.action in ['list', 'retrieve']:
            return ClassSessionReadSerializer
        return ClassSessionSerializer

    def get_queryset(self):
        user = self.request.user
        if not user.is_authenticated:
            return ClassSession.objects.none()
        if user.user_type == UserType.ADMIN:
            return ClassSession.objects.all()
        if user.user_type == UserType.TEACHER:
            return ClassSession.objects.filter(subject__teacher__user=user)
        return ClassSession.objects.none()

    def _ensure_teacher_owns_subject(self, subject):
        user = self.request.user
        if user.user_type == UserType.TEACHER and subject.teacher.user != user:
            raise PermissionDenied('You can only manage sessions for your own subjects.')

    def perform_create(self, serializer):
        self._ensure_teacher_owns_subject(serializer.validated_data['subject'])
        serializer.save()

    def perform_update(self, serializer):
        self._ensure_teacher_owns_subject(
            serializer.validated_data.get('subject', serializer.instance.subject)
        )
        serializer.save()
    
    def list(self, request, *args, **kwargs):
        self.queryset = apply_sorting(request, self.get_queryset(), self)
        return super().list(request, *args, **kwargs)
    
    @action(detail=False, methods=['get'], url_path='export_excel')
    def export_excel(self, request):
        """Export class sessions to Excel with filtering, searching, sorting"""
        queryset = self.get_queryset()
        
        filterset = self.filterset_class(request.GET, queryset=queryset)
        queryset = filterset.qs
        
        search_query = request.query_params.get('search', '')
        if search_query:
            from django.db.models import Q
            queryset = queryset.filter(
                Q(class_name__icontains=search_query) |
                Q(subject__name__icontains=search_query) |
                Q(subject__code__icontains=search_query)
            )
        
        ordering = request.query_params.get('ordering', '-date')
        if ordering:
            queryset = queryset.order_by(ordering)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Classes"
        
        headers = ['Class Name', 'Subject', 'Code', 'Date', 'Start Time', 'End Time', 'Teacher']
        ws.append(headers)
        
        from openpyxl.styles import Font, PatternFill
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        for session in queryset:
            teacher_name = ""
            if session.subject.teacher:
                teacher_name = f"{session.subject.teacher.user.first_name} {session.subject.teacher.user.last_name}".strip()
            
            ws.append([
                session.class_name,
                session.subject.name,
                session.subject.code,
                session.date.strftime('%Y-%m-%d') if session.date else "",
                session.start_time.strftime('%H:%M') if session.start_time else "",
                session.end_time.strftime('%H:%M') if session.end_time else "",
                teacher_name,
            ])
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 25
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="classes_{__import__("datetime").datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        return response
