from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from datetime import datetime
from django.db.models import Q
from academics.models import Enrollment
from ..models import Attendance
from core.utils.custom_perms import IsClientUser


class AdminReportsViewSet(viewsets.ViewSet):
    permission_classes = [IsClientUser]

    def _parse_date(self, date_str):
        try:
            return datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            return None

    def _get_teacher_name(self, teacher):
        first_name = (teacher.first_name or '').strip()
        last_name = (teacher.last_name or '').strip()
        full_name = f"{first_name} {last_name}".strip()
        return full_name if full_name else teacher.user.email

    def _build_session_report(self, attendance, queryset):
        session_id = str(attendance.class_session.id)
        date = attendance.marked_at.date().isoformat()
        
        total_enrolled = Enrollment.objects.filter(
            subject=attendance.class_session.subject
        ).count()
        
        session_attendances = queryset.filter(
            class_session=attendance.class_session,
            marked_at__date=date
        )
        
        present = session_attendances.filter(status='PRESENT').count()
        absent = session_attendances.filter(status='ABSENT').count()
        rate = (present / total_enrolled * 100) if total_enrolled > 0 else 0
        
        teacher = attendance.class_session.subject.teacher
        
        return {
            'session_id': session_id,
            'class_name': attendance.class_session.class_name,
            'subject_code': attendance.class_session.subject.code,
            'date': date,
            'teacher_name': self._get_teacher_name(teacher),
            'teacher_id': str(teacher.user_id),
            'department': attendance.student.department or 'Unknown',
            'total': total_enrolled,
            'present': present,
            'absent': absent,
            'attendance_rate': round(rate, 2)
        }

    @action(detail=False, methods=['get'])
    def sessions(self, request):
        try:
            start_date_str = request.query_params.get('start_date')
            end_date_str = request.query_params.get('end_date')
            department_filter = request.query_params.get('department')
            teacher_id_filter = request.query_params.get('teacher_id')
            status_filter = request.query_params.get('status')
            search_query = request.query_params.get('search', '').strip()
            sort_by = request.query_params.get('sort_by', 'date')
            sort_order = request.query_params.get('sort_order', 'desc')
            page = int(request.query_params.get('page', 1))
            page_size = int(request.query_params.get('page_size', 10))
            
            queryset = Attendance.objects.select_related(
                'class_session',
                'class_session__subject',
                'class_session__subject__teacher',
                'class_session__subject__teacher__user',
                'student',
                'student__user'
            )
            
            if start_date_str:
                start_date = self._parse_date(start_date_str)
                if start_date:
                    queryset = queryset.filter(marked_at__date__gte=start_date)
            
            if end_date_str:
                end_date = self._parse_date(end_date_str)
                if end_date:
                    queryset = queryset.filter(marked_at__date__lte=end_date)
            
            if department_filter:
                # Filter by student's department, not teacher's
                queryset = queryset.filter(student__department=department_filter)
            
            if teacher_id_filter:
                queryset = queryset.filter(class_session__subject__teacher__user_id=teacher_id_filter)
            
            if status_filter and status_filter in ['PRESENT', 'ABSENT']:
                queryset = queryset.filter(status=status_filter)
            
            if search_query:
                queryset = queryset.filter(Q(class_session__class_name__icontains=search_query))
            
            reports_data = []
            seen_keys = set()
            
            for attendance in queryset.order_by('-marked_at'):
                key = f"{attendance.class_session.id}-{attendance.marked_at.date().isoformat()}"
                
                if key not in seen_keys:
                    seen_keys.add(key)
                    reports_data.append(self._build_session_report(attendance, queryset))
            
            reverse_sort = sort_order.lower() == 'desc'
            if sort_by == 'class':
                reports_data.sort(key=lambda x: x['class_name'], reverse=reverse_sort)
            elif sort_by == 'rate':
                reports_data.sort(key=lambda x: x['attendance_rate'], reverse=reverse_sort)
            elif sort_by == 'present':
                reports_data.sort(key=lambda x: x['present'], reverse=reverse_sort)
            elif sort_by == 'department':
                reports_data.sort(key=lambda x: x['department'], reverse=reverse_sort)
            else:
                reports_data.sort(key=lambda x: x['date'], reverse=reverse_sort)
            
            if reports_data:
                avg_rate = sum(r['attendance_rate'] for r in reports_data) / len(reports_data)
                best_rate = max(r['attendance_rate'] for r in reports_data)
                worst_rate = min(r['attendance_rate'] for r in reports_data)
                total_present = sum(r['present'] for r in reports_data)
                total_absent = sum(r['absent'] for r in reports_data)
            else:
                avg_rate = best_rate = worst_rate = 0
                total_present = total_absent = 0
            
            total_count = len(reports_data)
            total_pages = (total_count + page_size - 1) // page_size
            start_idx = (page - 1) * page_size
            end_idx = start_idx + page_size
            
            return Response({
                'summary': {
                    'total_sessions': total_count,
                    'average_rate': round(avg_rate, 2),
                    'best_rate': round(best_rate, 2),
                    'worst_rate': round(worst_rate, 2),
                    'total_present': total_present,
                    'total_absent': total_absent,
                },
                'pagination': {
                    'current_page': page,
                    'page_size': page_size,
                    'total_count': total_count,
                    'total_pages': total_pages,
                },
                'sessions': reports_data[start_idx:end_idx]
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {'detail': f'Error fetching reports: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from io import BytesIO
            from django.http import HttpResponse
            
            start_date_str = request.query_params.get('start_date')
            end_date_str = request.query_params.get('end_date')
            department_filter = request.query_params.get('department')
            teacher_id_filter = request.query_params.get('teacher_id')
            status_filter = request.query_params.get('status')
            search_query = request.query_params.get('search', '').strip()
            sort_by = request.query_params.get('sort_by', 'date')
            sort_order = request.query_params.get('sort_order', 'desc')
            
            queryset = Attendance.objects.select_related(
                'class_session',
                'class_session__subject',
                'class_session__subject__teacher',
                'class_session__subject__teacher__user',
                'student',
                'student__user'
            )
            
            if start_date_str:
                start_date = self._parse_date(start_date_str)
                if start_date:
                    queryset = queryset.filter(marked_at__date__gte=start_date)
            
            if end_date_str:
                end_date = self._parse_date(end_date_str)
                if end_date:
                    queryset = queryset.filter(marked_at__date__lte=end_date)
            
            if department_filter:
                # Filter by student's department, not teacher's
                queryset = queryset.filter(student__department=department_filter)
            
            if teacher_id_filter:
                queryset = queryset.filter(class_session__subject__teacher__user_id=teacher_id_filter)
            
            if status_filter and status_filter in ['PRESENT', 'ABSENT']:
                queryset = queryset.filter(status=status_filter)
            
            if search_query:
                queryset = queryset.filter(Q(class_session__class_name__icontains=search_query))
            
            reports_data = []
            seen_keys = set()
            
            for attendance in queryset.order_by('-marked_at'):
                key = f"{attendance.class_session.id}-{attendance.marked_at.date().isoformat()}"
                
                if key not in seen_keys:
                    seen_keys.add(key)
                    reports_data.append(self._build_session_report(attendance, queryset))
            
            reverse_sort = sort_order.lower() == 'desc'
            if sort_by == 'class':
                reports_data.sort(key=lambda x: x['class_name'], reverse=reverse_sort)
            elif sort_by == 'rate':
                reports_data.sort(key=lambda x: x['attendance_rate'], reverse=reverse_sort)
            elif sort_by == 'present':
                reports_data.sort(key=lambda x: x['present'], reverse=reverse_sort)
            elif sort_by == 'department':
                reports_data.sort(key=lambda x: x['department'], reverse=reverse_sort)
            else:
                reports_data.sort(key=lambda x: x['date'], reverse=reverse_sort)
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Reports"
            
            headers = ['Date', 'Class', 'Subject Code', 'Teacher', 'Department', 'Total', 'Present', 'Absent', 'Rate (%)']
            ws.append(headers)
            
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            for row in reports_data:
                ws.append([
                    row['date'],
                    row['class_name'],
                    row['subject_code'],
                    row['teacher_name'],
                    row['department'],
                    row['total'],
                    row['present'],
                    row['absent'],
                    f"{row['attendance_rate']:.1f}"
                ])
            
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 10
            ws.column_dimensions['G'].width = 10
            ws.column_dimensions['H'].width = 10
            ws.column_dimensions['I'].width = 12
            
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="admin_reports_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
            
            return response

        except Exception as e:
            return Response(
                {'detail': f'Error exporting reports: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
