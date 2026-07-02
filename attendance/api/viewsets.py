from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from django.http import HttpResponse
from accounts.models import UserType
from academics.models import Enrollment, ClassSession
from ..models import FaceData, FaceEmbedding, Attendance, AttendanceLog, AttendanceSession
from ..filters import FaceDataFilter, FaceEmbeddingFilter, AttendanceFilter, AttendanceLogFilter
from rest_framework import serializers
from .serializers import (
    FaceDataSerializer, AttendanceSerializer, AttendanceReadSerializer, AttendanceLogSerializer,
    AttendanceMarkRequestSerializer, AttendanceSelfMarkRequestSerializer, AttendanceMarkResponseSerializer, SessionSummarySerializer, 
    SessionEndResponseSerializer, AttendanceSessionSerializer,
    StartSessionRequestSerializer, EndSessionRequestSerializer
)
from core.utils.custom_perms import IsClientUser
from core.utils.sort import apply_sorting
from ..ml_client import process_attendance, MLServiceError
from django.utils import timezone
import json
from drf_spectacular.utils import extend_schema
from django.db.models import Count, Q
from datetime import datetime


class FaceDataViewSet(viewsets.ModelViewSet):
    """
    Manage student face enrollment data.
    Only returns enrollment status (actual embeddings are never exposed via API).
    """
    queryset = FaceData.objects.all()
    serializer_class = FaceDataSerializer
    permission_classes = [IsClientUser]
    filterset_class = FaceDataFilter
    search_fields = ['student__user__email', 'student__roll_number']
    
    # Sorting configuration
    ordering_fields = ['id', 'student__user__email', 'student__roll_number', 'is_enrolled', 'total_photos_registered', 'registration_confidence', 'created_at', 'updated_at']
    ordering = ['-created_at']
    SORT_MAPPING = {
        'id': 'id',
        'student_id': 'student__id',
        'student_email': 'student__user__email',
        'student_roll_number': 'student__roll_number',
        'is_enrolled': 'is_enrolled',
        'total_photos_registered': 'total_photos_registered',
        'registration_confidence': 'registration_confidence',
        'created_at': 'created_at',
        'updated_at': 'updated_at',
    }
    
    def list(self, request, *args, **kwargs):
        self.queryset = apply_sorting(request, self.get_queryset(), self)
        return super().list(request, *args, **kwargs)
    
    @extend_schema(
        responses={200: FaceDataSerializer},
        description="Get current user's face enrollment status including registered photos count, registration confidence, and enrollment completion status."
    )
    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def my_enrollment_status(self, request):
        """Get current user's face enrollment status"""
        try:
            face_data = FaceData.objects.get(student__user=request.user)
            serializer = self.get_serializer(face_data)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except FaceData.DoesNotExist:
            return Response(
                {'detail': 'Face data not found for this user'},
                status=status.HTTP_404_NOT_FOUND
            )



class AttendanceViewSet(viewsets.ModelViewSet):
    """
    Handle attendance records. Supports:
    - Viewing attendance history
    - Filtering by student, class session, date range
    - Admin viewing all attendance
    - Students viewing their own attendance
    - Marking attendance via face recognition
    """
    permission_classes = [IsAuthenticated]
    parser_classes = (JSONParser, MultiPartParser, FormParser)
    filterset_class = AttendanceFilter
    search_fields = ['student__user__email', 'student__roll_number', 'class_session__class_name']
    
    # Sorting configuration
    ordering_fields = ['id', 'student__user__email', 'student__roll_number', 'class_session__class_name', 'class_session__date', 'status', 'marked_at', 'created_at', 'updated_at']
    ordering = ['-marked_at']
    SORT_MAPPING = {
        'id': 'id',
        'student_id': 'student__id',
        'student_email': 'student__user__email',
        'student_roll_number': 'student__roll_number',
        'class_session_id': 'class_session__id',
        'class_session_name': 'class_session__class_name',
        'class_date': 'class_session__date',
        'status': 'status',
        'marked_at': 'marked_at',
        'created_at': 'created_at',
        'updated_at': 'updated_at',
    }
    
    def list(self, request, *args, **kwargs):
        self.queryset = apply_sorting(request, self.get_queryset(), self)
        return super().list(request, *args, **kwargs)
    
    def get_queryset(self):
        """Filter attendance based on user role"""
        user = self.request.user
        if user.user_type == UserType.ADMIN:
            return Attendance.objects.all()
        elif user.user_type == UserType.STUDENT:
            return Attendance.objects.filter(student__user=user)
        elif user.user_type == UserType.TEACHER:
            # Teachers can see attendance for their subjects
            return Attendance.objects.filter(
                class_session__subject__teacher__user=user
            )
        return Attendance.objects.none()
    
    def get_serializer_class(self):
        if self.action == 'list' or self.action == 'retrieve':
            return AttendanceReadSerializer
        return AttendanceSerializer
    
    @extend_schema(
        request=AttendanceMarkRequestSerializer,
        responses={201: AttendanceMarkResponseSerializer, 200: AttendanceMarkResponseSerializer},
        description="Mark attendance using face recognition.\n\nProcess:\n1. Validates student has registered face\n2. Gets all enrolled students' embeddings\n3. Calls ML service to match face\n4. Creates Attendance record\n5. Creates AttendanceLog with verification metadata"
    )
    @action(detail=False, methods=['post'], permission_classes=[IsAuthenticated], parser_classes=[MultiPartParser, FormParser])
    def mark(self, request):
        """
        Mark attendance using face recognition.
        
        Request (multipart/form-data):
            - image: Face image file
            - class_session_id: UUID of the class session
            - latitude: GPS latitude (optional, but recommended)
            - longitude: GPS longitude (optional, but recommended)
            - liveness_passed: 'PASS', 'FAIL', or 'UNKNOWN'
            - timestamp_signed: Digitally signed timestamp from client
        
        Process:
            1. Validates student has registered face
            2. Gets all enrolled students' embeddings
            3. Calls ML service to match face
            4. Creates Attendance record
            5. Creates AttendanceLog with verification metadata
            6. Returns attendance result
        """
        if request.user.user_type != UserType.STUDENT:
            return Response(
                {'detail': 'Only students can mark attendance'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Validate required fields
        if 'image' not in request.FILES:
            return Response(
                {'detail': 'Image file is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        class_session_id = request.data.get('class_session_id')
        if not class_session_id:
            return Response(
                {'detail': 'class_session_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Get the class session
            class_session = ClassSession.objects.get(id=class_session_id)
        except ClassSession.DoesNotExist:
            return Response(
                {'detail': 'Class session not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Check if student is enrolled in this subject
        student_profile = request.user.studentprofile
        is_enrolled = Enrollment.objects.filter(
            student=student_profile,
            subject=class_session.subject
        ).exists()
        
        if not is_enrolled:
            return Response(
                {'detail': 'You are not enrolled in this subject'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Check if student has registered face
        try:
            student_face_data = FaceData.objects.get(student=student_profile)
            if not student_face_data.is_enrolled:
                return Response(
                    {'detail': 'You have not completed face registration yet. Please register first.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        except FaceData.DoesNotExist:
            return Response(
                {'detail': 'You have not registered your face yet. Please register first.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get all enrolled students' face embeddings in this subject — batch fetch
        # to avoid one DB query per student (N+1 problem).
        face_data_qs = (
            FaceData.objects
            .filter(
                student__enrollments__subject=class_session.subject,
                is_enrolled=True,
            )
            .select_related('student__user')
            .prefetch_related('embeddings')
        )

        stored_embeddings = []
        student_ids = []

        for face_data in face_data_qs:
            for embedding_record in face_data.embeddings.all():
                stored_embeddings.append(embedding_record.embedding)
                student_ids.append(str(face_data.student.user.id))
        
        if not stored_embeddings:
            return Response(
                {'detail': 'No enrolled students with registered faces found'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Call ML service for face matching
        image_file = request.FILES['image']
        
        try:
            ml_result = process_attendance(
                image_file,
                stored_embeddings,
                student_ids,
                session_id=str(class_session.id)
            )
        except MLServiceError as e:
            return Response(
                {'detail': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Extract ML result
        identified_student_id = ml_result.get('student_id')
        confidence = ml_result.get('confidence', 0.0)
        distance = ml_result.get('distance_to_nearest', float('inf'))
        ml_status = ml_result.get('status', 'unknown')
        
        # Determine attendance status based on face match
        if ml_status == 'identified' and identified_student_id:
            # Face was identified as current student
            attendance_status = 'PRESENT'
        else:
            # Face not identified or too far from database
            attendance_status = 'ABSENT'
        
        # Create or update Attendance record
        attendance, created = Attendance.objects.update_or_create(
            student=student_profile,
            class_session=class_session,
            defaults={'status': attendance_status}
        )
        
        # Extract optional verification data
        latitude = request.data.get('latitude')
        longitude = request.data.get('longitude')
        distance_from_classroom = request.data.get('distance_from_classroom')
        liveness_passed = request.data.get('liveness_passed', 'UNKNOWN')
        timestamp_signed = request.data.get('timestamp_signed')
        
        # Determine if attendance is suspicious
        is_suspicious = (
            ml_status != 'identified' or  # Face not recognized
            confidence < 0.3 or           # Low confidence match
            distance > 0.68               # High distance (above continuous detection threshold)
        )
        
        # Create AttendanceLog for verification layer
        AttendanceLog.objects.create(
            attendance=attendance,
            face_confidence=confidence,
            distance_to_nearest=distance,
            latitude=float(latitude) if latitude else None,
            longitude=float(longitude) if longitude else None,
            distance_from_classroom=float(distance_from_classroom) if distance_from_classroom else None,
            liveness_passed=liveness_passed,
            face_image_path=f"attendance/{student_profile.user.email}_{timezone.now().isoformat()}.jpg",
            is_suspicious=is_suspicious,
            timestamp_signed=timestamp_signed
        )
        
        return Response(
            {
                'attendance_id': str(attendance.id),
                'status': attendance_status,
                'marked_at': attendance.marked_at.isoformat(),
                'face_matched': ml_status == 'identified',
                'confidence': round(confidence, 4),
                'distance_to_nearest': round(distance, 6),
                'is_suspicious': is_suspicious,
                'message': f'Attendance marked as {attendance_status}'
            },
            status=status.HTTP_201_CREATED if created else status.HTTP_200_OK
        )
    
    @extend_schema(
        responses={200: AttendanceReadSerializer(many=True)},
        description="Get current student's attendance history. Only accessible to students. Returns list of attendance records with student, class session, and verification details."
    )
    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def my_attendance(self, request):
        """Get current student's attendance history"""
        if request.user.user_type != UserType.STUDENT:
            return Response(
                {'detail': 'Only students can access their attendance'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        attendances = Attendance.objects.filter(
            student__user=request.user
        ).order_by('-marked_at')
        
        serializer = AttendanceReadSerializer(attendances, many=True)
        return Response(serializer.data)
    
    @extend_schema(
        request=AttendanceSelfMarkRequestSerializer,
        responses={201: AttendanceMarkResponseSerializer, 200: AttendanceMarkResponseSerializer},
        description="Student self-marks attendance during active teacher session.\n\nRequirements:\n1. User must be a student\n2. Must be enrolled in the class\n3. Must have registered face\n4. Teacher must have started an active session for this class\n5. Class session date must be today\n6. Current time must be within class time window (with grace period)"
    )
    @action(detail=False, methods=['post'], permission_classes=[IsAuthenticated], parser_classes=[MultiPartParser, FormParser])
    def mark_self(self, request):
        """
        Student marks own attendance with face recognition.
        
        This endpoint is only available when teacher has started an active session.
        It validates:
        1. Student is enrolled in the class
        2. Student has registered face
        3. Teacher has an active session running
        4. Current time is within class time window
        
        Request (multipart/form-data):
            - image: Face image file
            - class_session_id: UUID of the class session
            - latitude: GPS latitude (optional)
            - longitude: GPS longitude (optional)
            - liveness_passed: 'PASS', 'FAIL', or 'UNKNOWN'
            - timestamp_signed: Digitally signed timestamp from client
        
        Returns: Attendance result with confidence and status
        """
        if request.user.user_type != UserType.STUDENT:
            return Response(
                {'detail': 'Only students can use this endpoint'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Validate required fields
        if 'image' not in request.FILES:
            return Response(
                {'detail': 'Image file is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        class_session_id = request.data.get('class_session_id')
        if not class_session_id:
            return Response(
                {'detail': 'class_session_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            class_session = ClassSession.objects.get(id=class_session_id)
        except ClassSession.DoesNotExist:
            return Response(
                {'detail': 'Class session not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Validate student is enrolled
        student_profile = request.user.studentprofile
        is_enrolled = Enrollment.objects.filter(
            student=student_profile,
            subject=class_session.subject
        ).exists()
        
        if not is_enrolled:
            return Response(
                {'detail': 'You are not enrolled in this subject'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Check if student has registered face
        try:
            student_face_data = FaceData.objects.get(student=student_profile)
            if not student_face_data.is_enrolled:
                return Response(
                    {'detail': 'You have not completed face registration yet'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        except FaceData.DoesNotExist:
            return Response(
                {'detail': 'You have not registered your face yet'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # ✅ CHECK: Teacher must have active session for this class
        active_session = AttendanceSession.objects.filter(
            class_session=class_session,
            ended_at__isnull=True  # Session still active
        ).first()
        
        if not active_session:
            return Response(
                {'detail': 'Teacher has not started attendance session for this class yet. Please wait for your teacher to start the session.'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # ✅ CHECK: Class session date must be today
        today = timezone.now().date()
        if class_session.date != today:
            return Response(
                {'detail': 'Attendance can only be marked for today\'s classes'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # ✅ CHECK: Current time must be within class time window (with 5 min grace)
        from datetime import timedelta
        now_time = timezone.now().time()
        grace_period = timedelta(minutes=5)
        end_time_with_grace = (
            timezone.datetime.combine(timezone.now().date(), class_session.end_time)
            + grace_period
        ).time()
        
        if not (class_session.start_time <= now_time <= end_time_with_grace):
            return Response(
                {'detail': f'Attendance can only be marked between {class_session.start_time} and {end_time_with_grace}'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # ✅ CHECK: Student hasn't already marked attendance for this CLASS (not just session)
        existing_attendance = Attendance.objects.filter(
            student=student_profile,
            class_session=class_session,  # Check whole class_session, not just attendance_session
        ).first()
        
        if existing_attendance:
            # If previously marked by student, block (no double marking)
            if existing_attendance.initiated_by == 'student':
                return Response(
                    {'detail': 'You have already marked attendance for this class'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            # If marked by teacher/admin, allow student to re-mark/verify themselves
            # This gives student a chance to verify with their own face capture
            logger.info(f"Student {student_profile.user.email} overriding {existing_attendance.initiated_by} attendance for {class_session.id}")
        
        # Get all enrolled students' face embeddings
        face_data_qs = (
            FaceData.objects
            .filter(
                student__enrollments__subject=class_session.subject,
                is_enrolled=True,
            )
            .select_related('student__user')
            .prefetch_related('embeddings')
        )

        stored_embeddings = []
        student_ids = []

        for face_data in face_data_qs:
            for embedding_record in face_data.embeddings.all():
                stored_embeddings.append(embedding_record.embedding)
                student_ids.append(str(face_data.student.user.id))
        
        if not stored_embeddings:
            return Response(
                {'detail': 'No enrolled students with registered faces found'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Call ML service
        image_file = request.FILES['image']
        
        try:
            ml_result = process_attendance(
                image_file,
                stored_embeddings,
                student_ids,
                session_id=str(active_session.id)
            )
        except MLServiceError as e:
            return Response(
                {'detail': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Extract ML result
        identified_student_id = ml_result.get('student_id')
        confidence = ml_result.get('confidence', 0.0)
        distance = ml_result.get('distance_to_nearest', float('inf'))
        ml_status = ml_result.get('status', 'unknown')
        
        # Determine attendance status based on ML confidence
        # Strict: Only mark PRESENT if own face identified with high confidence
        if (ml_status == 'identified' and 
            identified_student_id == str(request.user.id) and 
            confidence >= 0.85):  # High threshold for own face
            attendance_status = 'PRESENT'
            logger.info(f"PRESENT: Student {student_profile.user.email} face identified with {confidence:.2%} confidence")
        
        # Loose: Mark PRESENT if ANY face detected with moderate-high confidence (benefit of doubt)
        elif confidence >= 0.75:  # Moderate threshold
            attendance_status = 'PRESENT'
            logger.info(f"PRESENT: Face detected for {student_profile.user.email} with {confidence:.2%} confidence (not student-specific)")
        
        # Low confidence
        else:
            attendance_status = 'ABSENT'
            logger.info(f"ABSENT: Face confidence too low for {student_profile.user.email} (confidence={confidence:.2%})")
        
        # Create Attendance record (student-initiated)
        attendance, created = Attendance.objects.get_or_create(
            student=student_profile,
            class_session=class_session,
            attendance_session=active_session,
            defaults={
                'status': attendance_status,
                'initiated_by': 'student',
                'initiated_by_user': request.user,
                'frame_detected': timezone.now(),
                'detection_confidence': confidence,
                'attempt_count': 1
            }
        )
        
        # If not created (already exists), increment attempt count
        if not created:
            attendance.attempt_count += 1
            attendance.save()
        
        # Extract verification data
        latitude = request.data.get('latitude')
        longitude = request.data.get('longitude')
        distance_from_classroom = request.data.get('distance_from_classroom')
        liveness_passed = request.data.get('liveness_passed', 'UNKNOWN')
        timestamp_signed = request.data.get('timestamp_signed')
        
        # Determine if suspicious
        is_suspicious = (
            ml_status != 'identified' or
            confidence < 0.3 or
            distance > 0.68
        )
        
        # Create or update AttendanceLog
        AttendanceLog.objects.create(
            attendance=attendance,
            face_confidence=confidence,
            distance_to_nearest=distance,
            latitude=float(latitude) if latitude else None,
            longitude=float(longitude) if longitude else None,
            distance_from_classroom=float(distance_from_classroom) if distance_from_classroom else None,
            liveness_passed=liveness_passed,
            face_image_path=f"attendance/student/{student_profile.user.email}_{timezone.now().isoformat()}.jpg",
            is_suspicious=is_suspicious,
            timestamp_signed=timestamp_signed
        )
        
        # Broadcast attendance marked notification to all enrolled students
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        
        channel_layer = get_channel_layer()
        group_name = f'session_students_{class_session.id}'
        
        try:
            async_to_sync(channel_layer.group_send)(
                group_name,
                {
                    'type': 'attendance.marked',
                    'session_id': str(active_session.id),
                    'class_session_id': str(class_session.id),
                    'subject_code': class_session.subject.code,
                    'status': attendance_status,
                    'confidence': confidence,
                }
            )
        except Exception as e:
            print(f"Error broadcasting attendance_marked: {str(e)}")
        
        return Response(
            {
                'attendance_id': str(attendance.id),
                'status': attendance_status,
                'marked_at': attendance.marked_at.isoformat(),
                'face_matched': ml_status == 'identified',
                'confidence': round(confidence, 4),
                'distance_to_nearest': round(distance, 6),
                'is_suspicious': is_suspicious,
                'message': f'Attendance marked as {attendance_status}',
                'session_id': str(active_session.id)
            },
            status=status.HTTP_201_CREATED if created else status.HTTP_200_OK
        )
    
    @extend_schema(
        responses={200: AttendanceReadSerializer(many=True)},
        description="Get current student's attendance history. Only accessible to students. Returns list of attendance records with student, class session, and verification details."
    )
    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def my_attendance(self, request):
        """Get current student's attendance history"""
        if request.user.user_type != UserType.STUDENT:
            return Response(
                {'detail': 'Only students can access their attendance'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        attendances = Attendance.objects.filter(
            student__user=request.user
        ).order_by('-marked_at')
        
        serializer = AttendanceReadSerializer(attendances, many=True)
        return Response(serializer.data)
    
    @extend_schema(
        responses={200: AttendanceReadSerializer(many=True)},
        description="Get attendance records for a specific class session. Teacher only. Requires 'session_id' query parameter."
    )
    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def class_attendance(self, request):
        """Get attendance for a specific class session (teacher only)"""
        if request.user.user_type != UserType.TEACHER:
            return Response(
                {'detail': 'Only teachers can access class attendance'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        session_id = request.query_params.get('session_id')
        if not session_id:
            return Response(
                {'detail': 'session_id query parameter required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            session = ClassSession.objects.get(
                id=session_id,
                subject__teacher__user=request.user
            )
            attendances = Attendance.objects.filter(
                class_session=session
            ).order_by('student__roll_number')
            
            serializer = AttendanceReadSerializer(attendances, many=True)
            return Response(serializer.data)
        except ClassSession.DoesNotExist:
            return Response(
                {'detail': 'Class session not found or access denied'},
                status=status.HTTP_404_NOT_FOUND
            )
    
    @extend_schema(
        responses={200: SessionSummarySerializer},
        description="Get attendance summary for a class session including total students, present/absent counts, and attendance rate. Teacher only. Requires 'session_id' query parameter."
    )
    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def session_summary(self, request):
        """Get attendance summary for a class session (teacher only)"""
        if request.user.user_type != UserType.TEACHER:
            return Response(
                {'detail': 'Only teachers can access class attendance'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        session_id = request.query_params.get('session_id')
        if not session_id:
            return Response(
                {'detail': 'session_id query parameter required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            session = ClassSession.objects.get(
                id=session_id,
                subject__teacher__user=request.user
            )
            
            total_students = Enrollment.objects.filter(
                subject=session.subject
            ).count()
            
            present_count = Attendance.objects.filter(
                class_session=session,
                status='PRESENT'
            ).count()
            
            absent_count = Attendance.objects.filter(
                class_session=session,
                status='ABSENT'
            ).count()
            
            return Response({
                'session_id': str(session.id),
                'class_name': session.class_name,
                'date': session.date,
                'total_students': total_students,
                'present': present_count,
                'absent': absent_count,
                'attendance_rate': (present_count / total_students * 100) if total_students > 0 else 0
            })
        except ClassSession.DoesNotExist:
            return Response(
                {'detail': 'Class session not found or access denied'},
                status=status.HTTP_404_NOT_FOUND
            )

    @extend_schema(
        request=StartSessionRequestSerializer,
        responses={201: AttendanceSessionSerializer},
        description="Start real-time attendance session. Teacher initiates camera streaming for continuous face detection."
    )
    @action(detail=False, methods=['post'], permission_classes=[IsAuthenticated])
    def start_session(self, request, *args, **kwargs):
        """
        Start a new real-time attendance session.
        Only teachers can initiate sessions for their classes.
        
        Request (JSON):
            - class_session_id: UUID of the class session
        
        Response: AttendanceSession details with session ID for WebSocket connection
        """
        if request.user.user_type != UserType.TEACHER:
            return Response(
                {'detail': 'Only teachers can start attendance sessions'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        class_session_id = request.data.get('class_session_id')
        if not class_session_id:
            return Response(
                {'detail': 'class_session_id required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            class_session = ClassSession.objects.get(id=class_session_id)
            # Verify teacher teaches this class
            if class_session.subject.teacher.user != request.user:
                return Response(
                    {'detail': 'You do not teach this class'},
                    status=status.HTTP_403_FORBIDDEN
                )
        except ClassSession.DoesNotExist:
            return Response(
                {'detail': 'Class session not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Create new session
        from ..models import AttendanceSession
        session = AttendanceSession.objects.create(
            class_session=class_session,
            initiated_by=request.user,
            marked_students=[]
        )
        
        # Broadcast session started notification to all enrolled students
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        import json
        
        channel_layer = get_channel_layer()
        group_name = f'session_students_{class_session.id}'
        
        try:
            async_to_sync(channel_layer.group_send)(
                group_name,
                {
                    'type': 'session.started',
                    'session_id': str(session.id),
                    'class_session_id': str(class_session.id),
                    'subject_code': class_session.subject.code,
                    'subject_name': class_session.subject.name,
                    'template_id': str(class_session.template.id) if class_session.template else None,
                }
            )
        except Exception as e:
            print(f"Error broadcasting session_started: {str(e)}")
        
        serializer = AttendanceSessionSerializer(session)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @extend_schema(
        request=EndSessionRequestSerializer,
        responses={200: SessionEndResponseSerializer},
        description="End attendance session and auto-mark absent students"
    )
    @action(detail=False, methods=['post'], permission_classes=[IsAuthenticated])
    def end_session(self, request, *args, **kwargs):
        """
        End real-time attendance session and auto-mark absent students.
        
        Process:
            1. Closes the attendance session
            2. Gets all students enrolled in the class
            3. Marks as ABSENT any student NOT detected during session
            4. Returns summary statistics
        
        Request (JSON):
            - session_id: UUID of the AttendanceSession to close
        
        Response: Session summary with present/absent counts
        """
        session_id = request.data.get('session_id')
        if not session_id:
            return Response(
                {'detail': 'session_id required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        from ..models import AttendanceSession
        try:
            session = AttendanceSession.objects.get(id=session_id, ended_at=None)
        except AttendanceSession.DoesNotExist:
            return Response(
                {'detail': 'Session not found or already ended'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Verify requester is the teacher who started the session
        if session.initiated_by != request.user:
            return Response(
                {'detail': 'Only session initiator can end session'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Get all enrolled students in this class
        enrolled = Enrollment.objects.filter(subject=session.class_session.subject)
        
        # Mark as ABSENT if not detected
        marked_student_ids = session.marked_students
        for enrollment in enrolled:
            student_user_id = str(enrollment.student.user.id)
            if student_user_id not in marked_student_ids:
                Attendance.objects.get_or_create(
                    student=enrollment.student,
                    class_session=session.class_session,
                    attendance_session=session,
                    defaults={
                        'status': 'ABSENT',
                        'frame_detected': timezone.now(),
                        'detection_confidence': 0.0
                    }
                )
        
        # Close session
        session.ended_at = timezone.now()
        session.save()
        
        # Return summary
        present_count = Attendance.objects.filter(
            attendance_session=session,
            status='PRESENT'
        ).count()
        absent_count = Attendance.objects.filter(
            attendance_session=session,
            status='ABSENT'
        ).count()
        
        # Broadcast session ended notification to all enrolled students
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        
        channel_layer = get_channel_layer()
        group_name = f'session_students_{session.class_session.id}'
        
        try:
            async_to_sync(channel_layer.group_send)(
                group_name,
                {
                    'type': 'session.ended',
                    'session_id': str(session.id),
                    'class_session_id': str(session.class_session.id),
                    'subject_code': session.class_session.subject.code,
                    'marked_count': present_count,
                    'absent_count': absent_count,
                }
            )
        except Exception as e:
            print(f"Error broadcasting session_ended: {str(e)}")
        
        return Response({
            'session_id': str(session.id),
            'status': 'ended',
            'marked_present': present_count,
            'marked_absent': absent_count,
            'ended_at': session.ended_at.isoformat()
        }, status=status.HTTP_200_OK)

    @extend_schema(
        responses={200: serializers.Serializer()}
    )
    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def reports(self, request):
        """
        Get attendance reports with session-level statistics.
        
        Query Parameters:
            - start_date: YYYY-MM-DD (optional)
            - end_date: YYYY-MM-DD (optional)
            - status: PRESENT|ABSENT (optional)
            - search: Student name/roll number (optional)
            - sort_by: date|class|rate|present (default: date)
            - sort_order: asc|desc (default: desc)
            - page: Page number (default: 1)
            - page_size: Items per page (default: 10)
        
        Returns: Paginated sessions with attendance statistics
        """
        if request.user.user_type != UserType.TEACHER:
            return Response(
                {'detail': 'Only teachers can access reports'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Get query parameters
        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')
        status_filter = request.query_params.get('status')
        search_query = request.query_params.get('search', '').strip()
        sort_by = request.query_params.get('sort_by', 'date')
        sort_order = request.query_params.get('sort_order', 'desc')
        page = int(request.query_params.get('page', 1))
        page_size = int(request.query_params.get('page_size', 10))
        
        # Build queryset
        queryset = Attendance.objects.filter(
            class_session__subject__teacher__user=request.user
        ).select_related('class_session', 'class_session__subject', 'student', 'student__user')
        
        # Apply date filters
        if start_date_str:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                queryset = queryset.filter(marked_at__date__gte=start_date)
            except ValueError:
                return Response(
                    {'detail': 'Invalid start_date format. Use YYYY-MM-DD'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        if end_date_str:
            try:
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
                queryset = queryset.filter(marked_at__date__lte=end_date)
            except ValueError:
                return Response(
                    {'detail': 'Invalid end_date format. Use YYYY-MM-DD'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        # Apply status filter
        if status_filter and status_filter in ['PRESENT', 'ABSENT']:
            queryset = queryset.filter(status=status_filter)
        
        # Apply search filter (class session name)
        if search_query:
            queryset = queryset.filter(
                Q(class_session__class_name__icontains=search_query)
            )
        
        # Group by session and date
        reports_data = []
        seen_keys = set()
        
        for attendance in queryset.order_by('-marked_at'):
            session_id = str(attendance.class_session.id)
            date = attendance.marked_at.date().isoformat()
            key = f"{session_id}-{date}"
            
            if key not in seen_keys:
                seen_keys.add(key)
                
                # Get total enrolled students in this class
                total_enrolled = Enrollment.objects.filter(
                    subject=attendance.class_session.subject
                ).count()
                
                # Count attendance for this session/date
                session_attendances = queryset.filter(
                    class_session=attendance.class_session,
                    marked_at__date=date
                )
                
                present = session_attendances.filter(status='PRESENT').count()
                absent = session_attendances.filter(status='ABSENT').count()
                rate = (present / total_enrolled * 100) if total_enrolled > 0 else 0
                
                reports_data.append({
                    'session_id': session_id,
                    'class_name': attendance.class_session.class_name,
                    'subject_code': attendance.class_session.subject.code,
                    'date': date,
                    'total': total_enrolled,
                    'present': present,
                    'absent': absent,
                    'attendance_rate': round(rate, 2)
                })
        
        # Apply sorting
        reverse_sort = sort_order.lower() == 'desc'
        if sort_by == 'class':
            reports_data.sort(key=lambda x: x['class_name'], reverse=reverse_sort)
        elif sort_by == 'rate':
            reports_data.sort(key=lambda x: x['attendance_rate'], reverse=reverse_sort)
        elif sort_by == 'present':
            reports_data.sort(key=lambda x: x['present'], reverse=reverse_sort)
        else:  # date (default)
            reports_data.sort(key=lambda x: x['date'], reverse=reverse_sort)
        
        # Calculate summary stats (from all filtered data, not paginated)
        if reports_data:
            avg_rate = sum(r['attendance_rate'] for r in reports_data) / len(reports_data)
            best_rate = max(r['attendance_rate'] for r in reports_data)
            worst_rate = min(r['attendance_rate'] for r in reports_data)
            total_present = sum(r['present'] for r in reports_data)
            total_absent = sum(r['absent'] for r in reports_data)
        else:
            avg_rate = best_rate = worst_rate = 0
            total_present = total_absent = 0
        
        # Apply pagination
        total_count = len(reports_data)
        total_pages = (total_count + page_size - 1) // page_size
        start_idx = (page - 1) * page_size
        end_idx = start_idx + page_size
        paginated_data = reports_data[start_idx:end_idx]
        
        return Response({
            'summary': {
                'total_sessions': total_count,
                'average_rate': round(avg_rate, 2),
                'best_rate': round(best_rate, 2),
                'worst_rate': round(worst_rate, 2),
                'total_present': total_present,
                'total_absent': total_absent,
            },
            'pagination': {
                'current_page': page,
                'page_size': page_size,
                'total_count': total_count,
                'total_pages': total_pages,
            },
            'sessions': paginated_data
        })

    @extend_schema(
        responses={200: serializers.Serializer()}
    )
    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def export_excel(self, request):
        """
        Export attendance records as Excel file.
        
        Query Parameters (same as reports endpoint):
            - start_date: YYYY-MM-DD (optional)
            - end_date: YYYY-MM-DD (optional)
            - status: PRESENT|ABSENT (optional)
            - search: Student name/roll number (optional)
        
        Returns: Excel file download
        """
        if request.user.user_type != UserType.TEACHER:
            return Response(
                {'detail': 'Only teachers can export reports'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Get query parameters
        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')
        status_filter = request.query_params.get('status')
        search_query = request.query_params.get('search', '').strip()
        
        # Build queryset
        queryset = Attendance.objects.filter(
            class_session__subject__teacher__user=request.user
        ).select_related(
            'class_session', 'class_session__subject',
            'student', 'student__user',
            'verification_log'
        ).order_by('-marked_at')
        
        # Apply date filters
        if start_date_str:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                queryset = queryset.filter(marked_at__date__gte=start_date)
            except ValueError:
                return Response(
                    {'detail': 'Invalid start_date format. Use YYYY-MM-DD'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        if end_date_str:
            try:
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
                queryset = queryset.filter(marked_at__date__lte=end_date)
            except ValueError:
                return Response(
                    {'detail': 'Invalid end_date format. Use YYYY-MM-DD'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        # Apply status filter
        if status_filter and status_filter in ['PRESENT', 'ABSENT']:
            queryset = queryset.filter(status=status_filter)
        
        # Apply search filter
        if search_query:
            queryset = queryset.filter(
                Q(student__user__first_name__icontains=search_query) |
                Q(student__user__last_name__icontains=search_query) |
                Q(student__roll_number__icontains=search_query)
            )
        
        # Generate Excel
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from io import BytesIO
            from django.http import HttpResponse
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Attendance Report"
            
            # Define styles
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_alignment = Alignment(horizontal="center", vertical="center")
            
            # Write headers
            headers = ['Date', 'Time', 'Class', 'Subject Code', 'Student Name', 'Roll Number', 'Status', 'Confidence %', 'Liveness', 'Location Distance (m)']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = border
            
            # Write data rows
            for row_idx, attendance in enumerate(queryset, 2):
                confidence = ''
                if attendance.verification_log:
                    confidence = round(attendance.verification_log.face_confidence * 100, 1)
                
                liveness = 'N/A'
                if attendance.verification_log:
                    liveness = attendance.verification_log.liveness_passed
                
                distance = ''
                if attendance.verification_log and attendance.verification_log.distance_from_classroom:
                    distance = round(attendance.verification_log.distance_from_classroom, 1)
                
                row_data = [
                    attendance.marked_at.date().isoformat(),
                    attendance.marked_at.time().strftime('%H:%M:%S'),
                    attendance.class_session.class_name,
                    attendance.class_session.subject.code,
                    f"{attendance.student.user.first_name} {attendance.student.user.last_name}",
                    attendance.student.roll_number,
                    attendance.status,
                    confidence,
                    liveness,
                    distance,
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.value = value
                    cell.border = border
                    if col in [7, 8, 9, 10]:  # Numeric columns
                        cell.alignment = center_alignment
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 14
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 14
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 14
            ws.column_dimensions['I'].width = 14
            ws.column_dimensions['J'].width = 18
            
            # Save to BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            # Return as attachment
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="attendance_report_{timezone.now().date()}.xlsx"'
            return response
            
        except ImportError:
            return Response(
                {'detail': 'openpyxl library not installed. Please install it to export to Excel.'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @extend_schema(
    responses={200: serializers.Serializer()},
    description="WebSocket endpoint for real-time attendance streaming (Not a REST endpoint).\n\n"
                "WEBSOCKET PROTOCOL:\n"
                "Connection URL: ws://host/ws/attendance/stream/{session_id}/\n\n"
                "Flow:\n"
                "1. Teacher starts a session via POST /attendance/start_session/\n"
                "2. Connect to WebSocket with the returned session_id\n"
                "3. Send video frames continuously (binary data)\n"
                "4. Receive detection results in real-time\n"
                "5. End session via POST /attendance/end_session/\n\n"
                "Message Formats:\n"
                "INCOMING (Server → Client):\n"
                "{\n"
                "  'type': 'connection_established',\n"
                "  'status': 'connected',\n"
                "  'session_id': 'uuid',\n"
                "  'message': 'Ready to receive frames'\n"
                "}\n"
                "OR\n"
                "{\n"
                "  'type': 'frame_processed',\n"
                "  'newly_detected': [\n"
                "    {'student_id': 'uuid', 'student_email': 'email@example.com', "
                "'confidence': 0.92, 'detected_at': 'timestamp'}\n"
                "  ],\n"
                "  'timestamp': 'timestamp'\n"
                "}\n\n"
                "OUTGOING (Client → Server):\n"
                "Binary frame data (JPEG/PNG image bytes)\n\n"
                "Only the teacher who initiated the session can connect. "
                "Authentication is required via AuthMiddlewareStack.",
    tags=['Real-Time Attendance'],
    exclude=True  # Exclude from REST schema as it's WebSocket, not REST
    )
    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def websocket_stream(self, request):
        """WebSocket documentation endpoint (for schema only)"""
        pass

    
class AttendanceLogViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Verification logs for attendance. Read-only for auditing purposes.
    Admins can view all logs, teachers can view logs for their class sessions.
    """
    serializer_class = AttendanceLogSerializer
    permission_classes = [IsAuthenticated]
    filterset_class = AttendanceLogFilter
    search_fields = ['attendance__student__user__email', 'attendance__student__roll_number']
    
    # Sorting configuration
    ordering_fields = ['id', 'attendance__student__user__email', 'attendance__student__roll_number', 'attendance__class_session__class_name', 'face_confidence', 'distance_to_nearest', 'liveness_passed', 'created_at', 'updated_at']
    ordering = ['-created_at']
    SORT_MAPPING = {
        'id': 'id',
        'attendance_id': 'attendance__id',
        'student_id': 'attendance__student__id',
        'student_email': 'attendance__student__user__email',
        'student_roll_number': 'attendance__student__roll_number',
        'class_session_id': 'attendance__class_session__id',
        'class_session_name': 'attendance__class_session__class_name',
        'face_confidence': 'face_confidence',
        'distance_to_nearest': 'distance_to_nearest',
        'liveness_passed': 'liveness_passed',
        'created_at': 'created_at',
        'updated_at': 'updated_at',
    }
    
    def list(self, request, *args, **kwargs):
        self.queryset = apply_sorting(request, self.get_queryset(), self)
        return super().list(request, *args, **kwargs)
    
    def get_queryset(self):
        user = self.request.user
        if user.user_type == UserType.ADMIN:
            return AttendanceLog.objects.all()
        elif user.user_type == UserType.TEACHER:
            return AttendanceLog.objects.filter(
                attendance__class_session__subject__teacher__user=user
            )
        return AttendanceLog.objects.none()
    
    @extend_schema(
        responses={200: AttendanceLogSerializer(many=True)},
        description="Get all suspicious attendance records (admin only). Identifies attendance patterns flagged as suspicious based on face confidence, distance, and other metrics."
    )
    @action(detail=False, methods=['get'])
    def suspicious_activity(self, request):
        """Get all suspicious attendance logs (admin only)"""
        if request.user.user_type != UserType.ADMIN:
            return Response(
                {'detail': 'Only admins can view suspicious activity'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        logs = AttendanceLog.objects.filter(is_suspicious=True).order_by('-created_at')
        serializer = self.get_serializer(logs, many=True)
        return Response(serializer.data)
