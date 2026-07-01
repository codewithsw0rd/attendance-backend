from .serializers import *
from rest_framework.response import Response
from rest_framework import status
from rest_framework.viewsets import ModelViewSet
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.decorators import action
from django.db import transaction
from django.http import HttpResponse
from core.utils.custom_perms import IsClientUser
from core.utils.sort import apply_sorting
from ..filters import StudentProfileFilter, TeacherProfileFilter, AdminProfileFilter
from attendance.models import FaceData, FaceEmbedding
from attendance.ml_client import register_face_embedding, MLServiceError
from drf_spectacular.utils import extend_schema
import openpyxl
from io import BytesIO


class StudentViewSet(ModelViewSet):
    queryset = StudentProfile.objects.all()
    serializer_class = StudentProfileSerializer
    permission_classes = [IsClientUser]
    parser_classes = (MultiPartParser, FormParser)
    filterset_class = StudentProfileFilter
    search_fields = ['user__email', 'roll_number', 'first_name', 'last_name', 'department']
    
    # Sorting configuration
    ordering_fields = ['id', 'user__email', 'roll_number', 'department', 'year', 'first_name', 'last_name', 'created_at', 'updated_at']
    ordering = ['roll_number']
    SORT_MAPPING = {
        'id': 'id',
        'email': 'user__email',
        'roll_number': 'roll_number',
        'department': 'department',
        'year': 'year',
        'first_name': 'first_name',
        'last_name': 'last_name',
        'created_at': 'created_at',
        'updated_at': 'updated_at',
    }
    
    def get_serializer_class(self):
        """Use read serializer for list/retrieve, write serializer for create/update"""
        if self.action in ['list', 'retrieve']:
            return StudentProfileReadSerializer
        return StudentProfileSerializer
    
    def list(self, request, *args, **kwargs):
        self.queryset = apply_sorting(request, self.get_queryset(), self)
        return super().list(request, *args, **kwargs)
    
    @action(detail=False, methods=['get'], url_path='export_excel')
    def export_excel(self, request):
        """
        Export students to Excel file with filtering, searching, sorting support.
        Query parameters:
            - search: Search by email, roll_number, first_name, last_name
            - department: Filter by department
            - year: Filter by year
            - ordering: Sort field (e.g., 'first_name', '-created_at')
        """
        # Apply all the same filters/search/ordering as the list view
        queryset = self.get_queryset()
        
        # Apply filters
        filterset = self.filterset_class(request.GET, queryset=queryset)
        queryset = filterset.qs
        
        # Apply search
        search_query = request.query_params.get('search', '')
        if search_query:
            from django.db.models import Q
            queryset = queryset.filter(
                Q(user__email__icontains=search_query) |
                Q(roll_number__icontains=search_query) |
                Q(first_name__icontains=search_query) |
                Q(last_name__icontains=search_query)
            )
        
        # Apply ordering
        ordering = request.query_params.get('ordering', 'roll_number')
        if ordering:
            queryset = queryset.order_by(ordering)
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Students"
        
        # Add headers
        headers = ['Name', 'Email', 'Roll No', 'Department', 'Year', 'Status']
        ws.append(headers)
        
        # Style header row
        from openpyxl.styles import Font, PatternFill
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Add data rows
        for student in queryset:
            status_text = "Active" if student.user.is_active else "Inactive"
            ws.append([
                f"{student.first_name} {student.last_name}",
                student.user.email,
                student.roll_number,
                student.department or "",
                student.year or "",
                status_text,
            ])
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 12
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Return as file download
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="students_{__import__("datetime").datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        return response
    
    @extend_schema(
        request=StudentCreationRequestSerializer,
        responses={201: StudentCreationResponseSerializer},
        description="Create a new student and register their face photos in a single request.\n\nRequest must be multipart/form-data with email, first_name, last_name, roll_number, and images (1-5 face photos)."
    )
    @transaction.atomic
    def create(self, request, *args, **kwargs):
        """
        Create a new student and register their face photos in a single request.
        
        Request (multipart/form-data):
            - email: Student email (required)
            - first_name: Student first name (required)
            - last_name: Student last name (required)
            - roll_number: Student roll number (required)
            - images: List of face photos (upload multiple files with same field name 'images')
                      Minimum 1, Maximum 5 images
        
        Response:
            - Returns student profile + face enrollment status
        """
        request_data = request.data.copy()
        
        if 'user_type' not in request_data:
            request_data['user_type'] = UserType.STUDENT
        else:
            if request_data['user_type'] != UserType.STUDENT:
                return Response(
                    {'error': 'user_type must be STUDENT'},
                    status=status.HTTP_400_BAD_REQUEST
                )

        request_data['is_active'] = True
        
        # Extract face image files list
        face_images = request.FILES.getlist('images')
        
        # At least one face image is required
        if not face_images:
            return Response(
                {'error': 'At least one face image is required. Send multiple files with field name "images"'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Maximum 5 images
        if len(face_images) > 5:
            return Response(
                {'error': 'Maximum 5 face images allowed'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Create user and student profile
        user_serializer = CustomUserSerializer(data=request_data, context={'request': request})
        
        if not user_serializer.is_valid():
            return Response(user_serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        user_obj, access_token, refresh_token = user_serializer.save()
        student_profile = user_obj.studentprofile
        
        # Create FaceData record
        face_data, created = FaceData.objects.get_or_create(student=student_profile)
        
        # Process and register each face image
        try:
            for photo_number, image_file in enumerate(face_images, start=1):
                try:
                    # Call ML service to extract embedding and quality score
                    embedding, quality_score = register_face_embedding(image_file)

                    # Reject low-quality photos before storing — a bad embedding
                    # at enrollment causes poor matching for the lifetime of the account.
                    MIN_QUALITY = 0.35
                    if quality_score < MIN_QUALITY:
                        # IMPORTANT: @transaction.atomic only rolls back on exceptions,
                        # NOT on return statements. We must explicitly signal a rollback
                        # so the user/student created above is discarded.
                        transaction.set_rollback(True)
                        return Response(
                            {
                                'error': (
                                    f'Photo {photo_number} quality is too low '
                                    f'(score: {round(quality_score, 2)}, minimum: {MIN_QUALITY}). '
                                    'Please retake in better lighting with your face clearly visible '
                                    'and centred in the frame.'
                                )
                            },
                            status=status.HTTP_400_BAD_REQUEST
                        )

                    # Create FaceEmbedding record
                    FaceEmbedding.objects.create(
                        face_data=face_data,
                        embedding=embedding,
                        photo_number=photo_number,
                        quality_score=quality_score
                    )

                except MLServiceError as e:
                    # Roll back the user/student creation so the email can be reused.
                    transaction.set_rollback(True)
                    return Response(
                        {'error': f'Failed to process image {photo_number}: {str(e)}'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
            
            # Update FaceData with enrollment status
            total_photos = len(face_images)
            face_data.total_photos_registered = total_photos
            
            # Calculate average registration_confidence from all quality scores
            all_embeddings = FaceEmbedding.objects.filter(face_data=face_data)
            embedding_count = all_embeddings.count()
            if embedding_count > 0:
                avg_confidence = sum(e.quality_score for e in all_embeddings) / embedding_count
            else:
                avg_confidence = 0.0
            face_data.registration_confidence = avg_confidence
            
            # Mark as enrolled if 5 photos provided
            if total_photos >= 5:
                face_data.is_enrolled = True
            
            face_data.save()
            
            # Prepare response with student and face data
            student_response = StudentProfileReadSerializer(student_profile).data
            student_response['face_enrollment'] = {
                'total_photos_registered': face_data.total_photos_registered,
                'registration_confidence': round(face_data.registration_confidence, 4),
                'is_enrolled': face_data.is_enrolled,
                'message': 'Face registration completed' if total_photos == 5 else f'Face registration in progress ({total_photos}/5 photos)'
            }
            student_response['access_token'] = access_token
            student_response['refresh_token'] = refresh_token
            
            return Response(
                student_response,
                status=status.HTTP_201_CREATED
            )
        
        except Exception as e:
            # Roll back any DB writes before returning the 500.
            transaction.set_rollback(True)
            return Response(
                {'error': f'Error processing face images: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
class TeacherViewSet(ModelViewSet):
    queryset = TeacherProfile.objects.all()
    serializer_class = TeacherProfileSerializer
    permission_classes = [IsClientUser]
    parser_classes = (MultiPartParser, FormParser)
    filterset_class = TeacherProfileFilter
    search_fields = ['user__email', 'employee_id', 'first_name', 'last_name', 'department']
    
    # Sorting configuration
    ordering_fields = ['id', 'user__email', 'employee_id', 'department', 'first_name', 'last_name', 'created_at', 'updated_at']
    ordering = ['employee_id']
    SORT_MAPPING = {
        'id': 'id',
        'email': 'user__email',
        'employee_id': 'employee_id',
        'department': 'department',
        'first_name': 'first_name',
        'last_name': 'last_name',
        'created_at': 'created_at',
        'updated_at': 'updated_at',
    }
    
    def list(self, request, *args, **kwargs):
        self.queryset = apply_sorting(request, self.get_queryset(), self)
        return super().list(request, *args, **kwargs)
    
    @extend_schema(
        request=TeacherCreationRequestSerializer,
        responses={201: TeacherCreationResponseSerializer},
        description="Create a new teacher account.\n\nRequest fields: email, password, first_name, last_name, employee_id, and optional department, phone_no, address"
    )
    @transaction.atomic
    def create(self, request, *args, **kwargs):
        request_data = request.data.copy()
        
        if 'user_type' not in request_data:
            request_data['user_type'] = UserType.TEACHER
        else:
            if request_data['user_type'] != UserType.TEACHER:
                return Response(
                    {'error': 'user_type must be TEACHER'},
                    status=status.HTTP_400_BAD_REQUEST
                )

        request_data['is_active'] = True
        user_serializer = CustomUserSerializer(data=request_data, context={'request': request})
        
        if not user_serializer.is_valid():
            return Response(user_serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        user_obj, access_token, refresh_token = user_serializer.save()
        
        return Response(
            TeacherProfileReadSerializer(user_obj.teacherprofile).data,
            status=status.HTTP_201_CREATED
        )
        
class AdminViewSet(ModelViewSet):
    queryset = AdminProfile.objects.all()
    serializer_class = AdminProfileSerializer
    parser_classes = (MultiPartParser, FormParser)
    filterset_class = AdminProfileFilter
    search_fields = ['user__email', 'first_name', 'last_name']
    
    # Sorting configuration
    ordering_fields = ['id', 'user__email', 'first_name', 'last_name', 'created_at', 'updated_at']
    ordering = ['created_at']
    SORT_MAPPING = {
        'id': 'id',
        'email': 'user__email',
        'first_name': 'first_name',
        'last_name': 'last_name',
        'created_at': 'created_at',
        'updated_at': 'updated_at',
    }
    
    def list(self, request, *args, **kwargs):
        self.queryset = apply_sorting(request, self.get_queryset(), self)
        return super().list(request, *args, **kwargs)
    
    @extend_schema(
        request=AdminCreationRequestSerializer,
        responses={201: AdminCreationResponseSerializer},
        description="Create a new admin account. Maximum 3 admin users allowed.\n\nRequest fields: email, password, first_name, last_name, and optional phone_no, address"
    )
    @transaction.atomic
    def create(self, request, *args, **kwargs):
        request_data = request.data.copy()
        
        if CustomUser.objects.filter(user_type=UserType.ADMIN).count() >= 3:
            return Response(
                {'error': 'Only 3 admin users are allowed'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if 'user_type' not in request_data:
            request_data['user_type'] = UserType.ADMIN
        else:
            if request_data['user_type'] != UserType.ADMIN:
                return Response(
                    {'error': 'user_type must be ADMIN'},
                    status=status.HTTP_400_BAD_REQUEST
                )
                
        request_data['is_active'] = True
        user_serializer = CustomUserSerializer(data=request_data, context={'request': request})
        
        if not user_serializer.is_valid():
            return Response(user_serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        user_obj, _, _ = user_serializer.save()
        
        return Response(
            AdminProfileReadSerializer(user_obj.adminprofile).data,
            status=status.HTTP_201_CREATED
        )
        
    